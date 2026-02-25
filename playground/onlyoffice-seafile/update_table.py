import requests
from openpyxl import load_workbook
import os

TOKEN = ''  # API-token пользователя
SEAFILE_URL = ''  # Ссылка на сервис
REPO_ID = ''  # API библиотеки
PATH = '/test.xlsx'  # Путь до скачиваемого файла
LOCAL_PATH = 'download_test.xlsx'  # Имя файла после скачивания


def update_excel_in_seafile(token, seafile_url, repo_id, path, cell, value):
    headers = {"Authorization": f"Token {token}"}

    # Получение ссылки на скачивание
    api_url = f"{seafile_url}/api2/repos/{repo_id}/file/?p={path}"
    response = requests.get(api_url, headers=headers)
    response.raise_for_status()

    download_link = response.json()

    # Скачиваем временно
    temp_file = "temp.xlsx"
    file_response = requests.get(download_link)
    file_response.raise_for_status()

    with open(temp_file, "wb") as f:
        f.write(file_response.content)

    # Редактируем таблицу
    wb = load_workbook(temp_file)
    ws = wb[wb.sheetnames[0]]
    ws[cell] = value
    wb.save(temp_file)

    # Получаем upload-link
    upload_url_api = f"{seafile_url}/api2/repos/{repo_id}/upload-link/"
    upload_link_resp = requests.get(upload_url_api, headers=headers)
    upload_link_resp.raise_for_status()
    upload_link = upload_link_resp.json()

    # Загружаем обновлённый файл обратно
    with open(temp_file, "rb") as f:
        files = {
            'file': (os.path.basename(path), f),
            'parent_dir': (None, os.path.dirname(path)),
            'replace': (None, '1')  # Заменить существующий
        }

        upload_resp = requests.post(upload_link, headers=headers, files=files)
        upload_resp.raise_for_status()

    os.remove(temp_file)
    print("Файл обновлён в Seafile")


if __name__ == "__main__":
    update_excel_in_seafile(TOKEN, SEAFILE_URL, REPO_ID, PATH,
                            cell='A1', value='100')
