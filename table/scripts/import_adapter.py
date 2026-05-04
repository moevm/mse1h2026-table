import csv
import io
import os
from urllib.parse import quote

import requests
from openpyxl import Workbook, load_workbook

from scripts.utils import success, error

REQUEST_TIMEOUT = 60


def _build_dav_url(base_url, user, remote_path):
    base = f"{base_url.rstrip('/')}/remote.php/dav/files/{user}"
    parts = remote_path.lstrip("/").split("/")
    encoded_path = "/".join(quote(p) for p in parts)
    return f"{base}/{encoded_path}"


def _create_empty_xlsx_bytes(sheet_name=None):
    wb = Workbook()
    if sheet_name:
        wb.active.title = sheet_name
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _ensure_parent_dir(session, base_url, user, remote_path):
    parts = [p for p in remote_path.strip("/").split("/")[:-1] if p]
    if not parts:
        return
    base_dav = f"{base_url.rstrip('/')}/remote.php/dav/files/{user}"
    current = ""
    for part in parts:
        current += f"/{quote(part)}"
        url = f"{base_dav}{current}"
        propfind = session.request(
            "PROPFIND", url,
            headers={"Depth": "0"}, timeout=REQUEST_TIMEOUT,
        )
        if propfind.status_code == 404:
            mkcol = session.request("MKCOL", url, timeout=REQUEST_TIMEOUT)
            if mkcol.status_code != 201:
                error(
                    f"Failed to create directory (HTTP {mkcol.status_code}): "
                    f"{current}"
                )


def _read_csv(csv_path, separator, encoding, skip_columns):
    with open(csv_path, "r", encoding=encoding, newline="") as f:
        reader = csv.reader(f, delimiter=separator)
        try:
            raw_header = next(reader)
        except StopIteration:
            return [], []

        headers = raw_header[skip_columns:] if skip_columns else raw_header

        rows = []
        for raw_row in reader:
            row_values = raw_row[skip_columns:] if skip_columns else raw_row
            missing = len(headers) - len(row_values)
            if missing > 0:
                row_values = row_values + [""] * missing
            rows.append(dict(zip(headers, row_values)))

        return headers, rows


def _composite_key(row, key_cols):
    return tuple(
        "" if row.get(c) is None else str(row.get(c))
        for c in key_cols
    )


def _load_workbook_from_bytes(content, sheet_name):
    wb = load_workbook(io.BytesIO(content))
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            error(
                f"Sheet '{sheet_name}' not found in target. "
                f"Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    if ws.max_row == 1 and all(c.value is None for c in ws[1]):
        return wb, ws, []

    header_row = [
        "" if cell.value is None else str(cell.value)
        for cell in ws[1]
    ]
    return wb, ws, header_row


def _ensure_columns(ws, current_header, csv_headers):
    new_header = list(current_header)
    columns_added = []
    for h in csv_headers:
        if h not in new_header:
            new_header.append(h)
            columns_added.append(h)
            ws.cell(row=1, column=len(new_header), value=h)
    return new_header, columns_added


def _build_xlsx_index(ws, header, key_cols):
    try:
        key_indices = [header.index(c) for c in key_cols]
    except ValueError:
        return {}

    index = {}
    for r in range(2, ws.max_row + 1):
        key = tuple(
            "" if ws.cell(row=r, column=ki + 1).value is None
            else str(ws.cell(row=r, column=ki + 1).value)
            for ki in key_indices
        )
        if all(v == "" for v in key):
            continue
        index[key] = r
    return index


def _write_row(ws, row_idx, header, csv_row):
    for col_idx, col_name in enumerate(header, start=1):
        if col_name in csv_row:
            ws.cell(row=row_idx, column=col_idx, value=csv_row[col_name])


def import_run(args):
    """
    Обобщенный CSV -> Nextcloud xlsx адаптер с поддержкой upsert

    Читает CSV, сопоставляет колонки с существующей xlsx таблицей
    в Nextcloud по одному или более ключевым столбцам, и обновляет
    либо добавляет новые строки
    """
    key_cols = args.key
    if not key_cols:
        error(
            "--key is required "
            "(specify one or more times for composite key)"
        )

    if not os.path.exists(args.csv):
        error(f"CSV file not found: {args.csv}")

    if args.skip_columns < 0:
        error("--skip-columns must be >= 0")

    try:
        csv_headers, csv_rows = _read_csv(
            args.csv, args.separator, args.encoding, args.skip_columns
        )
    except UnicodeDecodeError as e:
        error(f"Encoding error reading CSV ({args.encoding}): {e}")
    except (OSError, csv.Error) as e:
        error(f"Failed to read CSV: {e}")

    if not csv_headers:
        error("CSV is empty (no header row)")

    missing_in_csv = [c for c in key_cols if c not in csv_headers]
    if missing_in_csv:
        error(
            f"Key column(s) not found in CSV header: {missing_in_csv}. "
            f"Available: {csv_headers}"
        )

    skipped_empty_key = 0
    valid_rows = []
    for row in csv_rows:
        key = _composite_key(row, key_cols)
        if all(v == "" for v in key):
            skipped_empty_key += 1
            continue
        valid_rows.append((key, row))

    with requests.Session() as session:
        session.auth = (args.username, args.password)
        dav_url = _build_dav_url(args.url, args.username, args.target)

        get_resp = session.get(dav_url, timeout=REQUEST_TIMEOUT)
        created = False
        if get_resp.status_code == 404:
            if not args.create_if_missing:
                error(
                    f"Target xlsx not found in Nextcloud: {args.target}. "
                    "Pass --create-if-missing to create it."
                )
            _ensure_parent_dir(
                session, args.url, args.username, args.target
            )
            content = _create_empty_xlsx_bytes(args.sheet)
            created = True
        elif get_resp.status_code >= 400:
            error(
                f"Failed to download target (HTTP {get_resp.status_code}): "
                f"{args.target}"
            )
        else:
            content = get_resp.content

        try:
            wb, ws, current_header = _load_workbook_from_bytes(
                content, args.sheet
            )
        except Exception as e:
            error(f"Failed to open xlsx: {e}")

        if not current_header:
            for i, h in enumerate(csv_headers, start=1):
                ws.cell(row=1, column=i, value=h)
            current_header = list(csv_headers)
            columns_added = []
        else:
            current_header, columns_added = _ensure_columns(
                ws, current_header, csv_headers
            )

        xlsx_index = _build_xlsx_index(ws, current_header, key_cols)

        rows_updated = 0
        rows_added = 0

        for key, row in valid_rows:
            if key in xlsx_index:
                _write_row(ws, xlsx_index[key], current_header, row)
                rows_updated += 1
            else:
                new_idx = ws.max_row + 1
                _write_row(ws, new_idx, current_header, row)
                xlsx_index[key] = new_idx
                rows_added += 1

        out = io.BytesIO()
        wb.save(out)

        put_resp = session.put(
            dav_url, data=out.getvalue(), timeout=REQUEST_TIMEOUT
        )
        if put_resp.status_code == 423:
            error(
                f"Target xlsx is locked (HTTP 423): {args.target}. "
                "Close the file in OnlyOffice and try again."
            )
        if put_resp.status_code not in (200, 201, 204):
            error(
                f"Failed to upload modified xlsx "
                f"(HTTP {put_resp.status_code}): {args.target}"
            )

    success({
        "target": args.target,
        "sheet": ws.title,
        "created": created,
        "csv_rows_read": len(csv_rows),
        "rows_updated": rows_updated,
        "rows_added": rows_added,
        "skipped_empty_key": skipped_empty_key,
        "key_columns": list(key_cols),
        "columns_added": columns_added,
    }, args.output)
